from openpyxl.utils import get_column_letter
from copy import copy

# FOR TYPE-HINTING
from openpyxl.worksheet.worksheet import Worksheet
from pandas import DataFrame


def incell_frame(worksheet: Worksheet,
                 dataframe: DataFrame,
                 col_range: tuple = (1, 0),
                 row_range: tuple = (1, 0),
                 num_str_cols: tuple = None,
                 float_cols: tuple = None,
                 skip_cols: tuple = None,
                 headers: bool = False):
    """Insert `dataframe` object into the Excel's working
    sheet - `worksheet` with flexibility to skipping
    columns (`skip_cols`), starting and stopping insertion process
    anywhere (`col_range`, `row_range`), inserting values as strings
    or floats or unchanged values (`num_str_cols`, `float_cols`) in
    case of having DataFrame columns like "ClientId"
    or "LoanApplicationId".

    Params:
        worksheet (openpyxl.worksheet.worksheet.Worksheet):
            Excel workbook's Worksheet object
        dataframe (pandas.DataFrame):
            DataFrame object of pandas
        col_range (tuple): From which to which column (index)
            to be inserted data from `dataframe`, by default it is
            (1, 0) which means it will start from the first column ("A")
            and will not stop until all the values from the related column
            of `dataframe` will be inserted (0 as the second element means
            not to stop until the very last column)
            [COLUMN INDEXES, NOT LETTERS]
        row_range (tuple): From which to which row to be
            inserted data from `dataframe`, by default it is
            (1, 0) which means it will start from the first row and
            will not stop until all the values from the `dataframe`
            will be inserted (0 as the second element means
            not to stop until the very last row)
        num_str_cols (tuple): Excel columns in which related values
            of `dataframe` should be inserted as unchanged string values
            EXAMPLE: ("A", "C", "AB") [COLUMN LETTERS, NOT INDEXES]
        float_cols (tuple): Excel columns in which related values
            of `dataframe` should be inserted as float values
            EXAMPLE: ("A", "C", "AB") [COLUMN LETTERS, NOT INDEXES]
        skip_cols (tuple): Excel columns to skip
            EXAMPLE: ("A", "C", "AB") [COLUMN LETTERS, NOT INDEXES]
        headers (bool): If True, DataFrame column headers will be inserted
            into the `worksheet`, too
    """
    # Starting to fill data from the DataFrame with
    # the first element of the tuple (starting column)
    # [FILLING PLAN: COLUMN AFTER COLUMN]
    col_index = col_range[0]
    for col in dataframe.columns:
        # If `col` related column to Excel exists in
        # the `skip_cols` tuple then skip it
        while get_column_letter(col_index) in skip_cols:
            col_index += 1

        # Starting row
        st_row = row_range[0]

        # If `headers` is True, insert headers as
        # the first row (first element value from `row_range`)
        if headers:
            worksheet[f'{get_column_letter(col_index)}{st_row}'] = str(col)

            st_row += 1

        # Start filling column cells starting from the first element
        # of `row_range` (starting row) with the related values
        # from the DataFrame
        for index, value in enumerate(dataframe[col], st_row):
            # In some cases, interpreting numeric values as string
            # is important, e.g., `LoanApplicationId`, `ClientId`
            # (values may as well start with the zeros)
            if get_column_letter(col_index) in num_str_cols:
                value = str(value)
            elif get_column_letter(col_index) in float_cols:
                value = float(value)

            worksheet[f'{get_column_letter(col_index)}{index}'] = value

            # If the row index is equal to the second
            # element of `row_range` (last row to fill)
            if (index + 1) == row_range[1]:
                break

        # If the column index is equal to the second
        # element of `col_range` (last column to fill)
        if (col_index + 1) == col_range[1]:
            break

        col_index += 1


def incell_style(cell_src, cell_dst):
    """Copy all cell style components/details from
    `cell_src` - the source cell to `cell_dst` - the destination
    cell.

    Params:
        cell_src (Worksheet[(str)]: Cell to copy the styles from
            [PLEASE MAKE SURE YOUR ARGUMENTS LOOK LIKE THE EXAMPLE BELLOW]
            worksheet["C12"]
        cell_dst (Worksheet[(str)]: Cell to copy the styles to
            [PLEASE MAKE SURE YOUR ARGUMENTS LOOK LIKE THE EXAMPLE BELLOW]
            worksheet["C12"]
    """
    cell_dst.font = copy(cell_src.font)
    cell_dst.border = copy(cell_src.border)
    cell_dst.fill = copy(cell_src.fill)
    cell_dst.number_format = copy(cell_src.number_format)
    cell_dst.protection = copy(cell_src.protection)
    cell_dst.alignment = copy(cell_src.alignment)
